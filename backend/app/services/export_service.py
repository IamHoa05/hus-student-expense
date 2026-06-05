# app/services/export_service.py

import io
from datetime import date, datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, extract, func
from typing import Optional

from ..models.transaction import Transaction, TransactionDetail, TransactionMedia
from ..models.financial import Budget
from ..models.category import Category, TransactionType
from ..config.database import get_db


class ExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_excel(
        self,
        user_id: int,
        month: Optional[int] = None,
        year: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> StreamingResponse:
        wb = Workbook()

        await self._sheet_transactions(wb, user_id, month, year, start_date, end_date)
        await self._sheet_budget(wb, user_id, month, year)

        # Xóa sheet mặc định
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = self._build_filename(month, year, start_date, end_date)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # -------------------------------------------------------------------------

    async def _sheet_transactions(self, wb, user_id, month, year, start_date, end_date):
        ws = wb.create_sheet("Giao dịch")

        headers = ["ID", "Ngày", "Loại", "Danh mục", "Số tiền", "Cửa hàng", "Ghi chú", "Thanh toán", "Địa điểm", "Nguồn"]
        self._write_header(ws, headers)

        # Build query
        stmt = (
            select(Transaction, Category.category_name, TransactionDetail)
            .join(Category, Transaction.category_id == Category.category_id, isouter=True)
            .outerjoin(TransactionDetail, Transaction.transaction_id == TransactionDetail.transaction_id)
            .where(Transaction.user_id == user_id)
        )
        stmt = self._apply_filter(stmt, Transaction.transaction_date, month, year, start_date, end_date)
        stmt = stmt.order_by(Transaction.transaction_date.desc())

        result = await self.db.execute(stmt)
        rows = result.all()

        for i, row in enumerate(rows, start=2):
            trans, cat_name, detail = row
            ws.append([
                trans.transaction_id,
                trans.transaction_date.strftime("%d/%m/%Y %H:%M"),
                "Chi" if trans.transaction_type == TransactionType.OUTFLOW else "Thu",
                cat_name or "",
                float(trans.total_amount),
                detail.store_name if detail else "",
                detail.note if detail else "",
                detail.payment_method if detail else "",
                detail.location if detail else "",
                trans.source.value
            ])

        self._auto_width(ws)

    async def _sheet_budget(self, wb, user_id, month, year):
        ws = wb.create_sheet("Ngân sách")

        headers = ["Danh mục", "Hạn mức", "Đã chi", "Còn lại", "% Sử dụng", "Ngưỡng cảnh báo", "Từ ngày", "Đến ngày", "Trạng thái"]
        self._write_header(ws, headers)

        today = date.today()
        _month = month or today.month
        _year = year or today.year

        stmt = (
            select(Budget, Category.category_name)
            .join(Category, Budget.category_id == Category.category_id)
            .where(
                and_(
                    Budget.user_id == user_id,
                    Budget.is_active == True,
                    extract("month", Budget.start_date) == _month,
                    extract("year", Budget.start_date) == _year
                )
            )
        )
        result = await self.db.execute(stmt)
        rows = result.all()

        for budget, cat_name in rows:
            spent = await self._calc_spent(user_id, budget.category_id, budget.start_date, budget.end_date)
            limit = float(budget.amount_limit)
            remaining = max(0, limit - spent)
            percentage = round((spent / limit) * 100, 1) if limit > 0 else 0

            ws.append([
                cat_name,
                limit,
                spent,
                remaining,
                percentage,
                float(budget.alert_threshold),
                budget.start_date.strftime("%d/%m/%Y"),
                budget.end_date.strftime("%d/%m/%Y"),
                "Vượt hạn mức" if percentage >= 100 else "Bình thường"
            ])

        self._auto_width(ws)

    # -------------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------------

    def _apply_filter(self, stmt, date_col, month, year, start_date, end_date):
        if start_date and end_date:
            return stmt.where(
                and_(
                    date_col >= datetime.combine(start_date, time.min),
                    date_col <= datetime.combine(end_date, time.max)
                )
            )
        if month and year:
            return stmt.where(
                and_(
                    extract("month", date_col) == month,
                    extract("year", date_col) == year
                )
            )
        if year:
            return stmt.where(extract("year", date_col) == year)
        return stmt

    async def _calc_spent(self, user_id, category_id, start, end) -> float:
        stmt = select(func.sum(Transaction.total_amount)).where(
            and_(
                Transaction.user_id == user_id,
                Transaction.category_id == category_id,
                Transaction.transaction_type == TransactionType.OUTFLOW,
                Transaction.transaction_date >= datetime.combine(start, time.min),
                Transaction.transaction_date <= datetime.combine(end, time.max)
            )
        )
        result = await self.db.execute(stmt)
        return float(result.scalar() or 0)

    def _write_header(self, ws, headers: list):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    def _build_filename(self, month, year, start_date, end_date) -> str:
        if start_date and end_date:
            return f"bao_cao_{start_date}_{end_date}.xlsx"
        if month and year:
            return f"bao_cao_{month:02d}_{year}.xlsx"
        if year:
            return f"bao_cao_{year}.xlsx"
        return f"bao_cao_{date.today()}.xlsx"


async def get_export_service(db: AsyncSession = Depends(get_db)):
    return ExportService(db)